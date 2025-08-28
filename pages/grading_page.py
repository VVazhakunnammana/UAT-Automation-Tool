import openpyxl
from openpyxl.styles import Alignment
import re
from datetime import datetime
from playwright.sync_api import Page
import time
from utils.excel_read import create_state_output_file
from utils.grading_model import create_grading_model

class GradingPage:
    """Page Object Model for the grading page."""
    
    def __init__(self, page: Page):
        self.page = page

    def grade_response(question, response, model=None):
        """
        Grade a response to a student question
        
        Args:
            question (str): The student's question
            response (str): The response to be graded
            model: The Gemini model (if None, creates a new one)
        
        Returns:
            str: Grading results including score and feedback
        """
        if model is None:
            model = create_grading_model()
        
        # Create the user prompt
        user_prompt = f"""You will receive a question and a response in the following format:

    <question>
    {question}
    </question>

    <response>
    {response}
    </response>

    Score the response on a scale of 0 to 100 based on the rubric. Output ONLY the numerical score."""
        
        # Get the evaluation
        chat = model.start_chat(history=[])
        evaluation = chat.send_message(user_prompt)
        
        return evaluation.text

    def extract_score(evaluation_text):
        """
        Extract the numerical score from the evaluation text
        
        Args:
            evaluation_text (str): The full evaluation text
        
        Returns:
            int: The numerical score, or None if not found
        """
        # First, try to find just a number (since AI should output only the score)
        # Look for patterns like "85", "92", etc.
        number_pattern = r'^(\d+)$'
        match = re.search(number_pattern, evaluation_text.strip())
        if match:
            score = int(match.group(1))
            if 0 <= score <= 100:
                return score
        
        # Fallback: Look for any number between 0-100 in the text
        fallback_pattern = r'\b(\d{1,2}|100)\b'
        match = re.search(fallback_pattern, evaluation_text)
        if match:
            score = int(match.group(1))
            if 0 <= score <= 100:
                return score
        
        # Legacy patterns (in case AI still outputs formatted text)
        score_patterns = [
            r'[Ss]core:\s*(\d+)/100',  # Score: 92/100
            r'[Ss]core:\s*(\d+)',       # Score: 92
            r'(\d+)/100',               # Just 92/100
        ]
        
        for pattern in score_patterns:
            match = re.search(pattern, evaluation_text, re.IGNORECASE)
            if match:
                score = int(match.group(1))
                if 0 <= score <= 100:
                    return score
        
        return None

    def read_mentor_configurations(self, config_file_path):
        """
        Reads mentor configurations from Real Estate AI Explainer.xlsx
        Returns: List of tuples [(state_name, mentor_url), ...]
        """
        mentors = []
        
        try:
            print(f"Reading mentor configurations from: {config_file_path}")
            # Use data_only=True to get calculated values instead of formulas
            workbook = openpyxl.load_workbook(config_file_path, data_only=True)
            
            sheet_name = "LLM-Url"
            if sheet_name not in workbook.sheetnames:
                print(f"Error: Sheet '{sheet_name}' not found")
                workbook.close()
                return mentors
                
            sheet = workbook[sheet_name]
            
            # Start from row 4 as specified
            row_count = 0
            empty_count = 0
            
            # Handle case where max_row might be None or very large
            # LIMITED TO 2 MENTORS FOR TESTING
            max_row = min(sheet.max_row or 1000, 5)# Only process first 2 mentors (rows 4-5)

            for row_num in range(2, max_row + 1):
                state_cell = sheet[f'A{row_num}']
                mentor_url_cell = sheet[f'B{row_num}']
                
                state_name = state_cell.value
                mentor_url = mentor_url_cell.value
                
                # Skip if either state or URL is empty
                if state_name and mentor_url:
                    # Clean the state name for file naming
                    state_name = str(state_name).strip()
                    mentor_url = str(mentor_url).strip()
                    
                    mentors.append((state_name, mentor_url))
                    row_count += 1
                    print(f"  Found mentor {row_count}: {state_name}")
                else:
                    empty_count += 1
                    
            workbook.close()
            
            print(f"\nTotal mentors found: {len(mentors)}")
            print(f"Empty rows skipped: {empty_count}")
            
            return mentors
            
        except Exception as e:
            print(f"Error reading mentor configurations: {str(e)}")
            return []
    
    def navigate_to_mentor_api(self, question, mentor_url):
        print("Navigating to Mentor API...")
        self.page.goto(mentor_url)

        # Wait for the page to load
        self.page.wait_for_load_state("networkidle")
        print("✓ Google loaded successfully")

        print("Searching for 'new'...")
        search_box = self.page.locator('textarea[data-testid="user-prompt-textarea"]')  # Text prompt input area
        search_box.fill(question)
                
        # # Press Enter to search
        search_box.press("Enter")
                
        # # Wait for search results to load
        self.page.wait_for_load_state("networkidle")
        print("✓ Response loaded successfully")


        # Find the search box and type "new"
        print("Searching for 'new'...")
        search_box = self.page.locator('textarea[data-testid="user-prompt-textarea"]')  # Text prompt input area
        search_box.fill("What steps should I take after obtaining my real estate license?")
    
        # Press Enter to search
        search_box.press("Enter")
    
        # Wait for search results to load
        self.page.wait_for_load_state("networkidle")
        print("✓ Response loaded successfully")

        # Click Copy button to Copy the response text
        copy_button = self.page.locator('[prop-events-value-onclick="handleCopyResponseBtnClick"]')
        copy_button.click()
        print("✓ Response text copied to clipboard")

        # # Get the response text from clipboard
        response_text = self.page.evaluate("navigator.clipboard.readText()")
                
        return response_text

    def read_questions_from_template(self, template_file_path):
        """
        Reads questions from the UAT Template Excel file
        Returns: List of questions
        """
        questions = []
        
        try:
            print(f"\nReading questions from: {template_file_path}")
            workbook = openpyxl.load_workbook(template_file_path, read_only=True)
            
            sheet_name = "Queries"
            if sheet_name not in workbook.sheetnames:
                print(f"Error: Sheet '{sheet_name}' not found")
                workbook.close()
                return questions
                
            sheet = workbook[sheet_name]
            
            # Read questions from column B, starting from row 2
            max_row = min(sheet.max_row or 100, 6)  # Limited to 6 questions (rows 2-7)
            for row_num in range(2, max_row + 1):
                prompt_cell = sheet[f'A{row_num}']
                if prompt_cell.value:
                    questions.append(str(prompt_cell.value).strip())
            
            workbook.close()
            print(f"Found {len(questions)} questions")
            
            return questions
            
        except Exception as e:
            print(f"Error reading questions: {str(e)}")
            return questions


    def process_mentor_questions(self, mentor_url, state_name, questions, model=None):
        """
        Processes all questions for a specific mentor and saves to state file
        
        Args:
            mentor_url (str): The mentor URL
            state_name (str): The state name for output file
            questions (list): List of questions to process
        """
        print(f"\n{'='*80}")
        print(f"PROCESSING MENTOR FOR: {state_name}")
        print(f"{'='*80}")
        print(f"Mentor URL: {mentor_url}")
        print(f"Total questions to process: {len(questions)}")
        
        try:
            # Create state output file
            workbook, sheet, file_path = create_state_output_file(state_name)
            ws = workbook.active

            processed_count = 0
            failed_count = 0
            current_row = 2  # Start from row 2 (after headers)

            # Process each question
            for idx, question in enumerate(questions, 1):
                print(f"\n[{idx}/{len(questions)}] Processing question {idx} for {state_name}")
                
                try:
                    # Get response from mentor
                    response = self.navigate_to_mentor_api(question, mentor_url)

                    # Write to Excel
                    sheet[f'A{current_row}'] = question
                    sheet[f'B{current_row}'] = response
                    sheet[f'C{current_row}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    sheet[f'D{current_row}'] = "Success"
                    
                    # Grade the response
                    evaluation = GradingPage.grade_response(str(question), str(response), model)
                    print(f"    AI Response: '{evaluation}'")

                    score = GradingPage.extract_score(evaluation)
                    print(f"    Extracted Score: {score}")
                    
                    # Write timestamp, status, and AI review score
                    current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    ws.cell(row=current_row, column=3, value=current_time)  # Column C: Timestamp
                    ws.cell(row=current_row, column=4, value="Success")     # Column D: Status
                    ws.cell(row=current_row, column=5, value=score if score is not None else "N/A")  # Column E: AI Review
                    
                    # Format timestamp column
                    ws.cell(row=current_row, column=3).alignment = Alignment(horizontal='center')

                    processed_count += 1
                    current_row += 1

                    # Save after each question to prevent data loss
                    workbook.save(file_path)
                    print(f"[OK] Question {idx} processed and saved")

                    
                except Exception as e:
                    # Log error but continue with next question
                    sheet[f'A{current_row}'] = question
                    sheet[f'B{current_row}'] = f"Error: {str(e)}"
                    sheet[f'C{current_row}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    sheet[f'D{current_row}'] = "Failed"
                    
                    failed_count += 1
                    current_row += 1
                    
                    workbook.save(file_path)
                    print(f"[FAILED] Question {idx} failed: {str(e)}")
                
                # Add delay between questions
                if idx < len(questions):
                    print("Waiting before next question...")
                    time.sleep(5)
            
            # Final save and close
            workbook.save(file_path)
            workbook.close()
            
            # Summary for this mentor
            print(f"\n{'='*60}")
            print(f"COMPLETED: {state_name}")
            print(f"{'='*60}")
            print(f"Processed: {processed_count}/{len(questions)}")
            print(f"Failed: {failed_count}")
            print(f"Output saved to: {file_path}")
            
            return processed_count, failed_count
            
        except Exception as e:
            print(f"Error processing mentor {state_name}: {str(e)}")
            return 0, len(questions)


