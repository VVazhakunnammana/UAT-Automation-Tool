import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from datetime import datetime
from pathlib import Path


def create_state_output_file(state_name):
    """
    Creates a new Excel file for the state with headers
    Returns: (workbook, sheet, file_path)
    """
    try:
        # Create output directory if it doesn't exist
        output_path = Path("output")
        output_path.mkdir(exist_ok=True)
        
        # Clean state name for file naming (remove special characters)
        clean_state_name = "".join(c for c in state_name if c.isalnum() or c in (' ', '-', '_')).rstrip()
        
        # Create filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{clean_state_name}_{timestamp}.xlsx"
        file_path = output_path / filename
        
        # Create new workbook
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = f"{state_name} Results"
        
        # Add headers
        sheet['A1'] = "Question"
        sheet['B1'] = "Response"
        sheet['C1'] = "Timestamp"
        sheet['D1'] = "Status"
        
        # Format headers
        for cell in sheet[1]:
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Adjust column widths
        sheet.column_dimensions['A'].width = 50
        sheet.column_dimensions['B'].width = 80
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 15

        # Check if Score and Evaluation columns already exist
        has_timestamp_column = False
        has_status_column = False
        has_ai_review_column = False
        has_rating_column = False
        has_bad_response_column = False
        has_notes_column = False
        has_fix_column = False
        has_ground_truth_column = False
        has_writer_column = False
        has_date_column = False
        
        # Check existing headers - look for any non-None value in the header row
        if sheet.cell(row=1, column=3).value is not None:
            has_timestamp_column = True
        if sheet.cell(row=1, column=4).value is not None:
            has_status_column = True
        if sheet.cell(row=1, column=5).value is not None:
            has_ai_review_column = True
        if sheet.cell(row=1, column=6).value is not None:
            has_rating_column = True
        if sheet.cell(row=1, column=7).value is not None:
            has_bad_response_column = True
        if sheet.cell(row=1, column=8).value is not None:
            has_notes_column = True
        if sheet.cell(row=1, column=9).value is not None:
            has_fix_column = True
        if sheet.cell(row=1, column=10).value is not None:
            has_ground_truth_column = True
        if sheet.cell(row=1, column=11).value is not None:
            has_writer_column = True
        if sheet.cell(row=1, column=12).value is not None:
            has_date_column = True

        print(f"  Column C header: '{sheet.cell(row=1, column=3).value}' (has_timestamp: {has_timestamp_column})")
        print(f"  Column D header: '{sheet.cell(row=1, column=4).value}' (has_status: {has_status_column})")
        print(f"  Column E header: '{sheet.cell(row=1, column=5).value}' (has_ai_review: {has_ai_review_column})")
        print(f"  Column F header: '{sheet.cell(row=1, column=6).value}' (has_rating: {has_rating_column})")
        print(f"  Column G header: '{sheet.cell(row=1, column=7).value}' (has_bad_response: {has_bad_response_column})")
        print(f"  Column H header: '{sheet.cell(row=1, column=8).value}' (has_notes: {has_notes_column})")
        print(f"  Column I header: '{sheet.cell(row=1, column=9).value}' (has_fix: {has_fix_column})")
        print(f"  Column J header: '{sheet.cell(row=1, column=10).value}' (has_ground_truth: {has_ground_truth_column})")
        print(f"  Column K header: '{sheet.cell(row=1, column=11).value}' (has_writer: {has_writer_column})")
        print(f"  Column L header: '{sheet.cell(row=1, column=12).value}' (has_date: {has_date_column})")

        # Add headers if they don't exist
        if not has_timestamp_column:
            sheet.cell(row=1, column=3, value="Timestamp")
            sheet.cell(row=1, column=3).font = Font(bold=True)
            sheet.cell(row=1, column=3).alignment = Alignment(horizontal='center')
            print("  Added Timestamp header to column C")
        
        if not has_status_column:
            sheet.cell(row=1, column=4, value="Status")
            sheet.cell(row=1, column=4).font = Font(bold=True)
            sheet.cell(row=1, column=4).alignment = Alignment(horizontal='center')
            print("  Added Status header to column D")
        
        if not has_ai_review_column:
            sheet.cell(row=1, column=5, value="AI Review")
            sheet.cell(row=1, column=5).font = Font(bold=True)
            sheet.cell(row=1, column=5).alignment = Alignment(horizontal='center')
            print("  Added AI Review header to column E")
        
        if not has_rating_column:
            sheet.cell(row=1, column=6, value="Rating")
            sheet.cell(row=1, column=6).font = Font(bold=True)
            sheet.cell(row=1, column=6).alignment = Alignment(horizontal='center')
            print("  Added Rating header to column F")
        
        if not has_bad_response_column:
            sheet.cell(row=1, column=7, value="If bad response, why?")
            sheet.cell(row=1, column=7).font = Font(bold=True)
            sheet.cell(row=1, column=7).alignment = Alignment(horizontal='center')
            print("  Added 'If bad response, why?' header to column G")
        
        if not has_notes_column:
            sheet.cell(row=1, column=8, value="Additional Notes")
            sheet.cell(row=1, column=8).font = Font(bold=True)
            sheet.cell(row=1, column=8).alignment = Alignment(horizontal='center')
            print("  Added Additional Notes header to column H")
        
        if not has_fix_column:
            sheet.cell(row=1, column=9, value="Fix?")
            sheet.cell(row=1, column=9).font = Font(bold=True)
            sheet.cell(row=1, column=9).alignment = Alignment(horizontal='center')
            print("  Added Fix? header to column I")
        
        if not has_ground_truth_column:
            sheet.cell(row=1, column=10, value="Ground Truth Version")
            sheet.cell(row=1, column=10).font = Font(bold=True)
            sheet.cell(row=1, column=10).alignment = Alignment(horizontal='center')
            print("  Added Ground Truth Version header to column J")
        
        if not has_writer_column:
            sheet.cell(row=1, column=11, value="Ground Truth Written By")
            sheet.cell(row=1, column=11).font = Font(bold=True)
            sheet.cell(row=1, column=11).alignment = Alignment(horizontal='center')
            print("  Added Ground Truth Written By header to column K")
        
        if not has_date_column:
            sheet.cell(row=1, column=12, value="Date")
            sheet.cell(row=1, column=12).font = Font(bold=True)
            sheet.cell(row=1, column=12).alignment = Alignment(horizontal='center')
            print("  Added Date header to column L")
        
        # Adjust column widths
        sheet.column_dimensions['C'].width = 20  # Timestamp column
        sheet.column_dimensions['D'].width = 15  # Status column
        sheet.column_dimensions['E'].width = 15  # AI Review column
        sheet.column_dimensions['F'].width = 15  # Rating column
        sheet.column_dimensions['G'].width = 25  # If bad response, why? column
        sheet.column_dimensions['H'].width = 25  # Additional Notes column
        sheet.column_dimensions['I'].width = 15  # Fix? column
        sheet.column_dimensions['J'].width = 25  # Ground Truth Version column
        sheet.column_dimensions['K'].width = 25  # Ground Truth Written By column
        sheet.column_dimensions['L'].width = 15  # Date column

        # Save initial file
        workbook.save(file_path)
        
        print(f"Created output file: {file_path}")
        return workbook, sheet, file_path
        
    except Exception as e:
        print(f"Error creating state output file: {str(e)}")
        raise
